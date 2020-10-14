from mendeley import Mendeley
import yaml
import os
# Get the DOI to look up
import argparse
import openpyxl
from openpyxl import Workbook
'''
u'Lecturer'
u'Researcher'
u'Professor > Associate Professor'
u'Professor'
u'Student  > Master'
u'Student  > Postgraduate'
u'Student  > Ph. D. Student'
u'Student  > Bachelor'
u'Student  > Graduate'
u'Student  > Doctoral Student'
u'Other'
u'Unspecified'
'''

titles=["Titles"]
Lecturer=[]
Researcher=[]
AssociateProfessor=[]
Professor=[]
Masters=[]
postgraduate=[]
PhDStudent=[]
Bachelor=[]
graduate=[]
DoctoralStudent=[]
others=[]
unspecified=[]

data=[]

config_file = 'config.yml'
config = {}
if os.path.isfile(config_file): 
    with open('config.yml') as f:
        config = yaml.load(f)
else:
    config['clientId'] = os.environ.get('MENDELEY_CLIENT_ID')
    config['clientSecret'] = os.environ.get('MENDELEY_CLIENT_SECRET')
men = Mendeley(config['clientId'], config['clientSecret'])
session = men.start_client_credentials_flow().authenticate()

def autometa(doi):
    doc = session.catalog.by_identifier(doi=doi, view='stats')
    titles.append(doc.title)
    data.append(doc.reader_count_by_academic_status)
    print ('"%s" has %s readers. \n' % (doc.title, doc.reader_count_by_academic_status))

    
path = "C:\\Users\\shiva\\Downloads\\STP\\doi.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
doil=[]
def iter_rows(sheet_obj):
    for row in sheet_obj.iter_rows():
        doil.append( [cell.value for cell in row][0])
iter_rows(sheet_obj)
doil=doil[1:]
for i in doil:
    try:
        autometa(str(i))
    except:
        print("cuased by ",i)
        print(str(i))
        print(doil.index(i))
doil.insert(0,"DOI's")
wb = Workbook()
ws = wb.active
for i in data:
    try:
        Lecturer.append(i.get(u'Lecturer',0))
        Researcher.append(i.get(u'Researcher',0))
        AssociateProfessor.append(i.get(u'Professor > Associate Professor',0))
        Professor.append(i.get(u'Professor',0))
        Masters.append(i.get(u'Student  > Master',0))
        postgraduate.append(i.get(u'Student  > Postgraduate',0))
        PhDStudent.append(i.get(u'Student  > Ph. D. Student',0))
        Bachelor.append(i.get(u'Student  > Bachelor',0))
        graduate.append(i.get(u'Student  > Graduate',0))
        DoctoralStudent.append(i.get(u'Student  > Doctoral Student',0))
        others.append(i.get(u'Other',0))
        unspecified.append(i.get(u'Unspecified',0))
    except:
        Lecturer.append(0)
        Researcher.append(0)
        AssociateProfessor.append(0)
        Professor.append(0)
        Masters.append(0)
        postgraduate.append(0)
        PhDStudent.append(0)
        Bachelor.append(0)
        graduate.append(0)
        DoctoralStudent.append(0)
        others.append(0)
        unspecified.append(0)
Lecturer.insert(0,"Lecturer")
Researcher.insert(0,"Researcher")
AssociateProfessor.insert(0,"AssociateProfessor")
Professor.insert(0,"Professor")
Masters.insert(0,"Masters")
postgraduate.insert(0,"Postgraduate")
PhDStudent.insert(0,"PhDStudent")
Bachelor.insert(0,"Bachelor")
graduate.insert(0,"Graduate")
DoctoralStudent.insert(0,"DoctoralStudent")
others.insert(0,"Others")
unspecified.insert(0,"Unspecified")
for row in zip(doil,titles,Lecturer,Researcher,AssociateProfessor,Professor,Masters,postgraduate,PhDStudent,Bachelor,graduate,DoctoralStudent,others,unspecified):
    ws.append(row)
    wb.save("columns.xlsx")
